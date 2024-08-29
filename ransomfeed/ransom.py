import sys
import importlib
from requests_html import HTMLSession
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from config.group_defaults import group_defaults
import googlemaps
import os
from dotenv import load_dotenv

load_dotenv()

def get_quarter(month):
    return (month - 1) // 3 + 1

def get_month_in_quarter(month):
    return (month - 1) % 3 + 1

def get_state(victim, config):
    google_maps_api_key = os.getenv('GOOGLE_MAPS_API_KEY')

    if not google_maps_api_key:
        print("Error: GOOGLE_MAPS_API_KEY not found in .env file")
        return 'Unknown', 'Unknown'

    gmaps = googlemaps.Client(key=google_maps_api_key)
    places_result = gmaps.places(query=f"{victim} {config.COUNTRY}")
    if places_result['status'] == 'OK' and places_result['results']:
        place = places_result['results'][0]
        place_details = gmaps.place(place_id=place['place_id'])
        if place_details['status'] == 'OK':
            address_components = place_details['result']['address_components']
            state = next((comp['long_name'] for comp in address_components if 'administrative_area_level_1' in comp['types']), 'Unknown')
            state = config.STATE_TRANSLATIONS.get(state, state)
            region = next((r for r, states in config.REGIONS.items() if state in states), 'Unknown')
            return region, state
    return 'Unknown', 'Unknown'

def main(country, ransom_id):
    try:
        config = importlib.import_module(f'config.{country.lower()}_config')
    except ImportError:
        print(f"Configuration for {country} not found.")
        sys.exit(1)

    s = HTMLSession()
    response = s.get(config.URL)
    table = response.html.xpath('//div/table/tbody/tr')
    data = []
    current_date = datetime.now()

    for row in reversed(table):
        cells = row.find('td')
        if int(cells[0].text) >= int(ransom_id):
            id = cells[0].text
            date_raw = cells[1].text
            victim = cells[2].text
            group = cells[3].text
            date = datetime.strptime(date_raw, '%Y-%m-%d %H:%M:%S')
            year = date.year
            month = date.month
            quarter = get_quarter(month)
            month_in_quarter = get_month_in_quarter(month)

            if config.API:
                region, state = get_state(victim, config)
            else:
                region, state = '', ''

            common_data = config.get_common_data(victim, region, state, month_in_quarter, quarter, year)

            if group in group_defaults:
                group_values = group_defaults[group]
                group_name = group_values[0]
                data.append(common_data + [
                    group_name, '',
                    f"L'azienda {victim} è stata colpita da un attacco ransomware sferrato dalla cybergang {group_name}",
                    '', '', 'Simone Felici',
                    f'https://ransomfeed.it/index.php?page=post_details&id_post={id}',
                    current_date, *group_values[1:]
                ])
            else:
                data.append(common_data + [
                    group, '',
                    f"L'azienda {victim} è stata colpita da un attacco ransomware sferrato dalla cybergang {group}",
                    '', '', 'Simone Felici',
                    f'https://ransomfeed.it/index.php?page=post_details&id_post={id}',
                    current_date, 'unknown', 'unknown', 'unknown', 'unknown', 'unknown',
                    'unknown', 'unknown', 'unknown', 'unknown', 'unknown',
                    'Data from Local System', 'unknown', 'unknown', 'Data Encrypted for Impact', 'NO', ''
                ])

    workbook = openpyxl.load_workbook('Brazil Threat Intelligence Report NEW VERSION.xlsx')
    worksheet = workbook['Details']

    # Crea un fill pattern rosso
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # Trova l'ultima riga utilizzata nel foglio
    last_row = worksheet.max_row

    for row_data in data:
        last_row += 1
        for col_num, cell_value in enumerate(row_data, start=1):
            cell = worksheet.cell(row=last_row, column=col_num, value=cell_value)
            cell.fill = red_fill
            if isinstance(cell_value, datetime):
                cell.number_format = 'mm/dd/yy'

    workbook.save(config.FILE)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python3 ransom.py <country> <ransom_id>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
